from jira import JIRA
import openpyxl
import pandas as pd
import tkinter as tk
from tkinter import ttk, scrolledtext, filedialog
from datetime import datetime
import os
import json
from openpyxl.styles import Font, PatternFill, Alignment

class UserInputWindow:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Jira 설정")
        self.root.geometry("600x500")
        
        # 설정 파일 경로
        self.config_file = 'jira_config.json'
        
        # 저장된 설정 불러오기
        self.load_config()
        
        self.create_widgets()
        
    def load_config(self):
        """저장된 설정 불러오기"""
        if os.path.exists(self.config_file):
            try:
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    self.config = json.load(f)
            except:
                self.config = {}
        else:
            self.config = {}
    
    def save_config(self):
        """설정 저장하기"""
        with open(self.config_file, 'w', encoding='utf-8') as f:
            json.dump(self.config, f, ensure_ascii=False, indent=2)
    
    def create_widgets(self):
        # 스타일 설정
        style = ttk.Style()
        style.configure('TLabel', padding=5)
        style.configure('TEntry', padding=5)
        style.configure('TButton', padding=5)
        
        # Email 입력
        ttk.Label(self.root, text="Email:").pack(fill='x', padx=5)
        self.email_entry = ttk.Entry(self.root, width=50)
        self.email_entry.pack(fill='x', padx=5)
        self.email_entry.insert(0, self.config.get('email', ''))
        
        # API Token 입력
        ttk.Label(self.root, text="API Token:").pack(fill='x', padx=5)
        self.token_entry = ttk.Entry(self.root, width=50)
        self.token_entry.pack(fill='x', padx=5)
        self.token_entry.insert(0, self.config.get('token', ''))
        
        # JQL Query 입력
        ttk.Label(self.root, text="JQL Query:").pack(fill='x', padx=5)
        self.query_text = scrolledtext.ScrolledText(self.root, wrap=tk.WORD, height=10)
        self.query_text.pack(fill='both', expand=True, padx=5, pady=5)
        self.query_text.insert('1.0', self.config.get('query', '''project = "CHUR" 
AND created >= "2024-11-04" 
AND created <= "2024-12-04" 
AND status IN (COMPLETE, "In Dev", "Known Issue", Resolved, Reopened, Open, "QA Review")'''))
        
        # 버튼 프레임
        button_frame = ttk.Frame(self.root)
        button_frame.pack(fill='x', padx=5, pady=10)
        
        # 저장 버튼
        ttk.Button(button_frame, text="저장", command=self.save_settings).pack(side='right', padx=5)
        
        # 취소 버튼
        ttk.Button(button_frame, text="취소", command=self.root.quit).pack(side='right', padx=5)
        
        # 설정 저장 체크박스
        self.save_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(button_frame, text="설정 저장", variable=self.save_var).pack(side='left', padx=5)
    
    def save_settings(self):
        """설정 저장 및 반환"""
        self.result = {
            'email': self.email_entry.get(),
            'token': self.token_entry.get(),
            'query': self.query_text.get('1.0', tk.END).strip()
        }
        
        # 설정 저장 체크박스가 선택된 경우
        if self.save_var.get():
            self.config = self.result.copy()
            self.save_config()
        
        self.root.quit()
    
    def get_settings(self):
        """설정값 반환"""
        self.root.mainloop()
        try:
            return self.result
        except AttributeError:
            return None

class ResultWindow:
    def __init__(self, matrix, df):
        self.root = tk.Toplevel()
        self.root.title("검색 결과")
        self.root.geometry("1000x600")
        
        # 노트북 (탭) 생성
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill='both', expand=True, padx=5, pady=5)
        
        # 매트릭스 탭
        self.create_matrix_tab(matrix)
        
        # 이슈 목록 탭
        self.create_issue_list_tab(df)
        
        # 확인 버튼
        ttk.Button(self.root, text="확인", command=self.root.destroy).pack(pady=5)
        
    def create_matrix_tab(self, matrix):
        matrix_frame = ttk.Frame(self.notebook)
        self.notebook.add(matrix_frame, text="집계 결과")
        
        # 트리뷰 생성
        tree = ttk.Treeview(matrix_frame)
        tree.pack(fill='both', expand=True, padx=5, pady=5)
        
        # 스크롤바 추가
        vsb = ttk.Scrollbar(matrix_frame, orient="vertical", command=tree.yview)
        vsb.pack(side='right', fill='y')
        tree.configure(yscrollcommand=vsb.set)
        
        # 열 설정
        headers = ['Status'] + list(matrix.columns)
        tree["columns"] = headers
        tree.column("#0", width=0, stretch=False)  # 숨김 열
        
        # 열 헤더 설정
        for header in headers:
            tree.heading(header, text=header)
            tree.column(header, width=100, anchor='center')
        
        # 데이터 추가
        for status in matrix.index:
            values = [status] + [int(matrix.loc[status, col]) for col in matrix.columns]
            tree.insert('', 'end', values=values)
            
    def create_issue_list_tab(self, df):
        issue_frame = ttk.Frame(self.notebook)
        self.notebook.add(issue_frame, text="이슈 목록")
        
        # 트리뷰 생성
        tree = ttk.Treeview(issue_frame)
        tree.pack(fill='both', expand=True, padx=5, pady=5)
        
        # 가로 스크롤바 추가
        hsb = ttk.Scrollbar(issue_frame, orient="horizontal", command=tree.xview)
        hsb.pack(side='bottom', fill='x')
        
        # 세로 스크롤바 추가
        vsb = ttk.Scrollbar(issue_frame, orient="vertical", command=tree.yview)
        vsb.pack(side='right', fill='y')
        
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        # DataFrame 컬럼명과 표시할 헤더 매핑
        column_mapping = {
            'key': 'Key',
            'summary': 'Summary',
            'status': 'Status',
            'priority': 'Priority',
            'created': 'Created',
            'assignee': 'Assignee'
        }
        
        # 열 설정
        headers = list(column_mapping.values())
        tree["columns"] = headers
        tree.column("#0", width=0, stretch=False)  # 숨김 열
        
        # 열 헤더 설정
        for header in headers:
            tree.heading(header, text=header)
            if header == 'Summary':
                tree.column(header, width=400, anchor='w')
            else:
                tree.column(header, width=100, anchor='center')
        
        # 데이터 추가
        df_columns = list(column_mapping.keys())  # DataFrame의 실제 컬럼명
        for _, row in df.iterrows():
            values = [row[col] for col in df_columns]
            tree.insert('', 'end', values=values)


    def show(self):
        self.root.grab_set()  # 모달 창으로 설정
        self.root.wait_window()  # 창이 닫힐 때까지 대기

def get_user_settings():
    """사용자 설정 입력 받기"""
    window = UserInputWindow()
    return window.get_settings()

def select_excel_file():
    """파일 선택 대화상자를 통해 Excel 파일 선택"""
    root = tk.Tk()
    root.withdraw()
    
    file_path = filedialog.askopenfilename(
        title="Excel 파일 선택",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
    )
    return file_path

def map_priority(priority):
    """한글 우선순위를 영문으로 매핑"""
    priority_map = {
        '블로커': 'Blocker',
        '심각': 'Critical',
        '주요': 'major',
        '사소': 'minor',
        '경미함': 'trivial'
    }
    return priority_map.get(priority.strip('[]'), priority)

def connect_to_jira(jira_server, email, api_token):
    """Jira 서버에 연결"""
    try:
        jira = JIRA(
            basic_auth=(email, api_token),
            server=jira_server
        )
        return jira
    except Exception as e:
        print(f"Jira 연결 실패: {str(e)}")
        return None

def standardize_status(status):
    """
    상태값을 표준화하는 함수
    '해결됨'과 'Resolved'를 동일하게 처리
    """
    status_mapping = {
        '해결됨': 'Resolved',
        '다시 열림': 'Reopened',
        # 필요에 따라 다른 상태값도 매핑 추가 가능
    }
    
    return status_mapping.get(status, status)

# get_jira_metrics 함수에서 이 유틸리티 함수를 호출하도록 수정
def get_jira_metrics(jira, jql_query):
    """Jira에서 이슈를 검색하고 매트릭스 형태로 집계"""
    try:
        print("\nJira에서 이슈 검색 중...")
        
        # 전체 이슈 가져오기 (페이징 처리)
        start_at = 0
        max_results = 100
        all_issues = []
        
        while True:
            issues = jira.search_issues(jql_query, startAt=start_at, maxResults=max_results)
            if len(issues) == 0:
                break
                
            all_issues.extend(issues)
            start_at += len(issues)
            
        print(f"총 {len(all_issues)}개의 이슈를 찾았습니다.")
        
        # 데이터 수집
        data = []
        for issue in all_issues:
            priority = issue.fields.priority.name
            if priority.startswith('[') and priority.endswith(']'):
                priority = priority[1:-1]
            
            # 상태값 표준화 적용
            status = standardize_status(str(issue.fields.status))
                
            data.append({
                'key': issue.key,
                'summary': issue.fields.summary,
                'status': status,  # 표준화된 상태값 사용
                'priority': map_priority(priority),
                'created': issue.fields.created[:10],
                'assignee': str(issue.fields.assignee) if issue.fields.assignee else 'Unassigned'
            })
            
        df = pd.DataFrame(data)
        
        # 상태별 이슈 수 확인 (디버깅 용도)
        unique_statuses = df['status'].unique()
        print(f"발견된 고유 상태값: {unique_statuses}")
        
        matrix = pd.crosstab(df['status'], df['priority'])
        
        status_order = [
            'COMPLETE',
            'Resolved',  # 표준화된 이름만 포함
            'OPEN',
            'Known Issue',
            'In Dev',
            'Reopened',
            'QA Review'
        ]
        
        priority_order = ['Blocker', 'Critical', 'major', 'minor', 'trivial']
        
        for status in status_order:
            if status not in matrix.index:
                matrix.loc[status] = 0
                
        for priority in priority_order:
            if priority not in matrix.columns:
                matrix[priority] = 0
                
        matrix = matrix.reindex(status_order)
        matrix = matrix.reindex(columns=priority_order)
        matrix['Total'] = matrix.sum(axis=1)
        matrix.loc['Total'] = matrix.sum()
        new_index = ['Total'] + [idx for idx in matrix.index if idx != 'Total']
        matrix = matrix.reindex(new_index)
        
        return matrix, df
        
    except Exception as e:
        print(f"\n이슈 검색 실패: {str(e)}")
        return None, None

def update_excel_matrix(file_path, matrix, df):
    """Excel 파일의 기존 시트에 매트릭스와 이슈 목록을 업데이트"""
    try:
        print("\n=== Excel 파일 업데이트 시작 ===")
        
        # 워크북 열기
        workbook = openpyxl.load_workbook(file_path)
        print(f"현재 시트 목록: {workbook.sheetnames}")
        
        # 1. 매트릭스 데이터 시트 처리
        matrix_sheet_name = "Total_issue"
        if matrix_sheet_name in workbook.sheetnames:
            print(f"기존 '{matrix_sheet_name}' 시트 데이터 업데이트")
            matrix_sheet = workbook[matrix_sheet_name]
            # 기존 데이터 삭제
            matrix_sheet.delete_rows(1, matrix_sheet.max_row)
        else:
            print(f"새로운 '{matrix_sheet_name}' 시트 생성")
            matrix_sheet = workbook.create_sheet(title=matrix_sheet_name, index=0)
        
        # 매트릭스 스타일 설정
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # 헤더 작성
        headers = ['Status'] + list(matrix.columns)
        for col, header in enumerate(headers, 1):
            cell = matrix_sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        # 매트릭스 데이터 입력
        for i, status in enumerate(matrix.index, 2):
            status_cell = matrix_sheet.cell(row=i, column=1)
            status_cell.value = status
            status_cell.font = header_font
            status_cell.alignment = Alignment(horizontal='left', vertical='center')
            
            for j, column in enumerate(matrix.columns, 2):
                cell = matrix_sheet.cell(row=i, column=j)
                cell.value = int(matrix.loc[status, column])
                cell.alignment = center_alignment
        
        # 매트릭스의 마지막 행 번호 확인
        last_row = len(matrix.index) + 1

        # QA 상태 분석 추가
        analysis = analyze_qa_status(df, matrix)
        if analysis:
            # 빈 행 추가
            matrix_sheet.cell(row=last_row + 2, column=1, value="").font = header_font
            
            # 분석 결과 헤더
            matrix_sheet.cell(row=last_row + 3, column=1, value="QA 현황 분석").font = header_font
            matrix_sheet.cell(row=last_row + 3, column=2, value=f"Status: {analysis['status'].replace('_', ' ')}").font = header_font
            
            current_row = last_row + 4
            
            # 위험 요소
            if analysis['points']['risks']:
                matrix_sheet.cell(row=current_row, column=1, value="위험 요소:").font = header_font
                current_row += 1
                for risk in analysis['points']['risks']:
                    matrix_sheet.cell(row=current_row, column=2, value=f"• {risk}")
                    current_row += 1
                current_row += 1
            
            # 긍정적 요소
            if analysis['points']['positives']:
                matrix_sheet.cell(row=current_row, column=1, value="긍정적 요소:").font = header_font
                current_row += 1
                for positive in analysis['points']['positives']:
                    matrix_sheet.cell(row=current_row, column=2, value=f"• {positive}")
                    current_row += 1
                current_row += 1
            
            # 필요 조치사항
            if analysis['points']['needs']:
                matrix_sheet.cell(row=current_row, column=1, value="필요 조치사항:").font = header_font
                current_row += 1
                for need in analysis['points']['needs']:
                    matrix_sheet.cell(row=current_row, column=2, value=f"• {need}")
                    current_row += 1
            
            # 셀 스타일 적용
            for row in range(last_row + 3, current_row):
                matrix_sheet.row_dimensions[row].height = 20
                for col in range(1, 7):  # A부터 F열까지
                    cell = matrix_sheet.cell(row=row, column=col)
                    if row == last_row + 3:  # 헤더 행
                        cell.fill = PatternFill(start_color=analysis['status_color'], 
                                              end_color=analysis['status_color'], 
                                              fill_type="solid")
                    cell.alignment = Alignment(vertical='center')

        # 열 너비 자동 조정
        for column in matrix_sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min((max_length + 2), 50)
            matrix_sheet.column_dimensions[column[0].column_letter].width = adjusted_width
            
            
        # 2. 이슈 목록 시트 처리
        issue_sheet_name = "Issue List"
        if issue_sheet_name in workbook.sheetnames:
            print(f"기존 '{issue_sheet_name}' 시트 데이터 업데이트")
            issue_sheet = workbook[issue_sheet_name]
            # 기존 데이터 삭제
            issue_sheet.delete_rows(1, issue_sheet.max_row)
        else:
            print(f"새로운 '{issue_sheet_name}' 시트 생성")
            issue_sheet = workbook.create_sheet(title=issue_sheet_name, index=1)
        
        # 이슈 목록 헤더
        issue_headers = ['Key', 'Summary', 'Status', 'Priority', 'Created', 'Assignee']
        for col, header in enumerate(issue_headers, 1):
            cell = issue_sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        # 이슈 데이터 입력
        for i, row in enumerate(df.itertuples(), 2):
            issue_sheet.cell(row=i, column=1, value=row.key)
            issue_sheet.cell(row=i, column=2, value=row.summary)
            issue_sheet.cell(row=i, column=3, value=row.status)
            issue_sheet.cell(row=i, column=4, value=row.priority)
            issue_sheet.cell(row=i, column=5, value=row.created)
            issue_sheet.cell(row=i, column=6, value=row.assignee)
        
        # 이슈 목록 시트 열 너비 조정
        for column in issue_sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # 최대 너비 제한
            issue_sheet.column_dimensions[column[0].column_letter].width = adjusted_width
            # QA Daily 시트 업데이트
        if 'QA Daily' in workbook.sheetnames:
            update_qa_daily(workbook, df)
        
        # 파일 저장
        print("\n변경사항 저장 중...")
        workbook.save(file_path)
        print(f"Excel 파일 업데이트 완료: {file_path}")
        print(f"최종 시트 목록: {workbook.sheetnames}")
        
    except Exception as e:
        print(f"\nExcel 파일 업데이트 실패: {str(e)}")
        import traceback
        traceback.print_exc()


def main():
    # 사용자 설정 입력 받기
    settings = get_user_settings()
    if not settings:
        print("프로그램을 종료합니다.")
        return
        
    JIRA_SERVER = "https://nzin-publisher-bts.atlassian.net"
    EMAIL = settings['email']
    API_TOKEN = settings['token']
    jql_query = settings['query']
    
    # Excel 파일 선택
    print("\nExcel 파일을 선택해주세요.")
    excel_path = select_excel_file()
    if not excel_path:
        print("파일이 선택되지 않았습니다. 프로그램을 종료합니다.")
        return
        
    # Jira 연결
    jira = connect_to_jira(JIRA_SERVER, EMAIL, API_TOKEN)
    if not jira:
        return
    
    # Jira 메트릭스 데이터 가져오기
    matrix, df = get_jira_metrics(jira, jql_query)
    if matrix is not None:
        # 결과 확인 출력
        print("\n상태별 이슈 분포:")
        print(df['status'].value_counts())
        
        # Excel 파일 업데이트
        update_excel_matrix(excel_path, matrix, df)
        
        # 결과 창 표시
        result_window = ResultWindow(matrix, df)
        result_window.show()

def update_qa_daily(workbook, df):
    """QA Daily 시트에 오늘 생성된 이슈 통계 업데이트"""
    try:
        today = datetime.now().strftime('%Y-%m-%d')
        qa_sheet = workbook['QA Daily']
        
        # 오늘 생성된 이슈 필터링
        today_issues = df[df['created'] == today]
        total_today = len(today_issues)
        
        # 우선순위별 카운트
        priority_counts = today_issues['priority'].value_counts()
        
        # 통계 문자열 생성
        stats_text = f"금일 생성된 이슈: {total_today}건\n"
        if total_today > 0:
            stats_text += "우선순위별 현황:\n"
            for priority, count in priority_counts.items():
                stats_text += f"- {priority}: {count}건\n"
        
        # QA Daily 시트 E4 셀에 업데이트
        qa_sheet['E4'] = stats_text
        
        # 셀 줄바꿈 설정
        qa_sheet['E4'].alignment = Alignment(wrapText=True)
        
        print(f"\nQA Daily 시트 업데이트 완료")
        print(stats_text)
        
    except Exception as e:
        print(f"QA Daily 시트 업데이트 실패: {str(e)}")

def update_qa_daily(workbook, df):
    """QA Daily 시트에 오늘 생성된 이슈 통계 업데이트"""
    try:
        today = datetime.now().strftime('%Y-%m-%d')
        qa_sheet = workbook['QA Daily']
        
        # 오늘 생성된 이슈 필터링
        today_issues = df[df['created'] == today]
        total_today = len(today_issues)
        
        # 우선순위별 카운트
        priority_counts = today_issues['priority'].value_counts()
        
        # 우선순위별 카운트를 지정된 셀에 입력
        priority_cells = {
            'Blocker': 'F73',
            'Critical': 'G73',
            'major': 'H73',
            'minor': 'I73',
            'trivial': 'J73'
        }
        
        # 모든 셀을 먼저 0으로 초기화
        for cell in priority_cells.values():
            qa_sheet[cell] = 0
            qa_sheet[cell].alignment = Alignment(horizontal='center', vertical='center')
        
        # 카운트된 우선순위별 이슈 수 입력
        for priority, count in priority_counts.items():
            if priority in priority_cells:
                qa_sheet[priority_cells[priority]] = count
                qa_sheet[priority_cells[priority]].alignment = Alignment(horizontal='center', vertical='center')
        
        # 통계 문자열 생성 (E4 셀용)
        stats_text = f"금일 생성된 이슈: {total_today}건\n"
        if total_today > 0:
            stats_text += "우선순위별 현황:\n"
            for priority, count in priority_counts.items():
                stats_text += f"- {priority}: {count}건\n"
        
        # E4 셀 업데이트
        qa_sheet['E4'] = stats_text
        qa_sheet['E4'].alignment = Alignment(wrapText=True, vertical='center')
        
        print(f"\nQA Daily 시트 업데이트 완료")
        print(stats_text)
        print("\n우선순위별 셀 입력:")
        for priority, cell in priority_cells.items():
            count = qa_sheet[cell].value
            print(f"{priority} ({cell}): {count}건")
        
    except Exception as e:
        print(f"QA Daily 시트 업데이트 실패: {str(e)}")
        import traceback
        traceback.print_exc()
        raise

# update_excel_matrix 함수의 마지막 부분에 다음 코드 추가 (파일 저장 직전)
        # QA Daily 시트 업데이트
        if 'QA Daily' in workbook.sheetnames:
            print("\nQA Daily 시트 업데이트 시작...")
            update_qa_daily(workbook, df)
        else:
            print("\nQA Daily 시트가 없습니다.")
        
        # 파일 저장
        print("\n변경사항 저장 중...")
        workbook.save(file_path)
        print(f"Excel 파일 업데이트 완료: {file_path}")
        print(f"최종 시트 목록: {workbook.sheetnames}")
        
    except Exception as e:
        print(f"\nExcel 파일 업데이트 실패: {str(e)}")
        import traceback
        traceback.print_exc()

# QA 상태 분석
        analysis = analyze_qa_status(df, matrix)
        if analysis:
            # 상태에 따른 배경색 설정
            status_colors = {
                'GOOD': 'C6EFCE',      # 연한 녹색
                'MEDIUM_RISK': 'FFEB9C',  # 연한 노란색
                'HIGH_RISK': 'FFC7CE'   # 연한 빨간색
            }
            
            # E5 셀에 분석 결과 표시
            status_text = f"■ QA 현황 분석 (Status: {analysis['status'].replace('_', ' ')})\n\n"
            
            if analysis['points']['risks']:
                status_text += "📌 위험 요소:\n"
                for risk in analysis['points']['risks']:
                    status_text += f"- {risk}\n"
                status_text += "\n"
            
            if analysis['points']['positives']:
                status_text += "✔️ 긍정적 요소:\n"
                for positive in analysis['points']['positives']:
                    status_text += f"- {positive}\n"
                status_text += "\n"
            
            if analysis['points']['needs']:
                status_text += "⚠️ 필요 조치사항:\n"
                for need in analysis['points']['needs']:
                    status_text += f"- {need}\n"
            
            qa_sheet['B5'] = status_text
            qa_sheet['B5'].alignment = Alignment(wrapText=True, vertical='top')
            
            # 배경색 설정
            qa_sheet['B5'].fill = PatternFill(start_color=status_colors[analysis['status']], 
                                            end_color=status_colors[analysis['status']], 
                                            fill_type="solid")

def analyze_qa_status(df, matrix):
    """QA 상태 분석"""
    try:
        # 상태별, 우선순위별 이슈 수 확인
        in_dev_issues = df[df['status'] == 'In dev']
        complete_issues = df[df['status'] == 'COMPLETE']
        in_dev_counts = in_dev_issues['priority'].value_counts()
        complete_counts = complete_issues['priority'].value_counts()

        # 상태 색상 정의
        status_colors = {
            'GOOD': 'C6EFCE',      # 연한 녹색
            'MEDIUM_RISK': 'FFEB9C',  # 연한 노란색
            'HIGH_RISK': 'FFC7CE'   # 연한 빨간색
        }

        # 분석 결과
        analysis = {
            'status': 'MEDIUM_RISK',  # 기본값을 MEDIUM_RISK로 설정
            'status_color': status_colors['MEDIUM_RISK'],  # 기본 색상 설정
            'points': {
                'risks': [],      # 위험 요소
                'positives': [],  # 긍정적인 요소
                'needs': []       # 필요한 조치사항
            }
        }

        # 1. Blocker 이슈 분석
        blocker_count = in_dev_counts.get('Blocker', 0)
        if blocker_count > 0:
            analysis['status'] = 'HIGH_RISK'
            analysis['status_color'] = status_colors['HIGH_RISK']
            analysis['points']['risks'].append(f"Blocker 이슈 {blocker_count}건이 In dev 상태")
            analysis['points']['needs'].append("Blocker 이슈 우선 처리 필요")

        # 2. Major/Minor 이슈 분석
        major_in_dev = in_dev_counts.get('major', 0)
        minor_in_dev = in_dev_counts.get('minor', 0)
        
        if major_in_dev > 40:
            analysis['points']['risks'].append(f"Major 이슈 {major_in_dev}건이 In dev 상태로 다수 누적")
            analysis['points']['needs'].append("Major 이슈 처리 속도 개선 필요")
        
        # 3. 긍정적 요소 분석
        major_complete = complete_counts.get('major', 0)
        minor_complete = complete_counts.get('minor', 0)

        if major_complete > 0 or minor_complete > 0:
            analysis['points']['positives'].append(
                f"Major 이슈 {major_complete}건, Minor 이슈 {minor_complete}건 COMPLETE 처리 완료"
            )

        if 'Critical' not in in_dev_counts:
            analysis['points']['positives'].append("Critical 이슈 없음")

        # 4. 전체적인 상태 분석
        total_in_dev = len(in_dev_issues)
        total_issues = len(df)
        in_dev_ratio = total_in_dev / total_issues if total_issues > 0 else 0

        if in_dev_ratio > 0.7:  # 70% 이상이 In dev
            analysis['points']['risks'].append(f"전체 이슈의 {in_dev_ratio:.1%}가 In dev 상태")
            analysis['points']['needs'].append("개발팀과 진행 상황 점검 필요")

        # 상태가 GOOD이 될 수 있는 조건
        if not analysis['points']['risks'] and blocker_count == 0:
            analysis['status'] = 'GOOD'
            analysis['status_color'] = status_colors['GOOD']
            if not analysis['points']['positives']:
                analysis['points']['positives'].append("위험 요소 없이 정상적으로 진행 중")

        return analysis

    except Exception as e:
        print(f"QA 상태 분석 실패: {str(e)}")
        return None

if __name__ == "__main__":
    main()